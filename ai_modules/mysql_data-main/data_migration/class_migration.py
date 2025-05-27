

# 엑셀 파일에서 '교과목구분' 열에서 전공선택, 전공필수 만 남기고 지움
# 엑셀 파일에서 '학과학년' 열에서 농생명과학과, 한옥건축학과, 농업시스템학과, 지역산업학과, 에코농산업벤처시스템학과 지움

# 에러 "엑셀에 적힌 학과 이름" : 딕셔너리에 해당 에러에 나타난 학과 이름의 key가 없는 것
# 에러 "실제 학과 이름" : 딕셔너리의 value 값에 대응하는 jbnu_department 테이블의 학과 이름이 존재하지 않음. 또는 jbnu_class 테이블 자체의 SQL 구문 오류


import mysql.connector
from openpyxl import load_workbook
from tkinter import filedialog
import copy
from dotenv import load_dotenv
import os

load_dotenv()  # .env 파일 자동 로드

api_key = os.getenv("OPENAI_API_KEY")
db_host = os.getenv("DB_HOST")
db_password = os.getenv("DB_PASSWORD")

host = db_host
port = '3311'
database = 'nll_second'
user = 'root'
password = db_password

connection = mysql.connector.connect(host=host, port=port, database=database, user=user, password=password)
cursor = connection.cursor()

wb = load_workbook(filedialog.askopenfilename())
ws = wb.active

# 학과 이름 딕셔너리. key는 엑셀에 적힌 이름, value는 변환될 실제 학과 이름
department_dic = {
    "간호" : "간호학과",
    "건축공" : "건축공학과",
    "경영" : "경영학과",
    "경제부(경제)" : "경제학부",
    "경제학부(경제)" : "경제학부",
    "고고문화인" : "고고문화인류학과",
    "고분자.나노공" : "고분자나노공학과",
    "고분자섬유나노공학부(고분자나노공학)" : "고분자나노공학과",
    "고분자섬유나노공학부(유기소재섬유공학)" : "고분자나노공학과",
    "공공인재학부" : "공공인재학부",
    "과학" : "과학학과",
    "과학교(물리)" : "과학교육학부",
    "과학교(생물)" : "과학교육학부",
    "과학교(지구과학)" : "과학교육학부",
    "과학교(화학)" : "과학교육학부",
    "교육" : "교육학과",
    "국어교육" : "국어교육과",
    "국어국문" : "국어국문학과",
    "국제이공학부(엔지니어링사이언스)" : "국제이공학부",
    "국제인문사회학부(지미카터국제협력)" : "국제인문사회학부",
    "기계공" : "기계공학과",
    "기계설계(기계설계)" : "기계설계공학부",
    "기계설계(나노바이오)" : "기계설계공학부",
    "기계시스템" : "기계시스템공학부",
    "정밀기계" : "기계시스템공학부",
    "응용기계" : "기계시스템공학부",
    "농경제유통학부(농업경제학)" : "농경제유통학부",
    "농경제유통학부(식품유통학)" : "농경제유통학부",
    "농생물학과" : "농생물학과",
    "도시공" : "도시공학과",
    "독어교육" : "독어교육과",
    "독일학" : "독일학과",
    "동물생명공학과" : "동물생명공학과",
    "동물자원과" : "동물자원과학과",
    "목재응용과학과" : "목재응용과학과",
    "무역" : "무역학과",
    "무용" : "무용학과",
    "문헌정보" : "문헌정보학과",
    "물리" : "물리학과",
    "미술" : "미술학과",
    "바이오(헬스케어기기)" : "바이오메디컬공학부",
    "바이오(메디컬AI)" : "바이오메디컬공학부",
    "바이오(헬스케어정보)" : "바이오메디컬공학부",
    "반도체과학기술학" : "반도체과학기술학과",
    "사학과" : "사학과",
    "사회" : "사회학과",
    "사회복지" : "사회복지학과",
    "산림환경과학과" : "산림환경과학과",
    "산림환경과학과(재)" : "산림환경과학과",
    "산업디자인학" : "산업디자인학과",
    "산업정보시스템공" : "산업정보시스템공학과",
    "생명공(생명자원소재)" : "생명공학부",
    "생명공(환경생명)" : "생명공학부",
    "생명과학부(분자생물학)" : "생명과학부(분자생물학)",
    "생명과학부(생명과학)" : "생명과학부(생명과학)",
    "생명자원융합학과" : "생명자원융합학과",
    "생물산업기계" : "생물산업기계공학과",
    "생물환경화학과" : "생물환경화학과",
    "생태조경디자인학과" : "생태조경디자인학과",
    "소프트웨어공" : "소프트웨어공학과",
    "소프트웨어공학과(재)" : "소프트웨어공학과",
    "수" : "수학과",
    "수의" : "수의학과",
    "수학교육" : "수학교육과",
    "스마트팜학과" : "스마트팜학과",
    "스페인.중남미학과" : "스페인.중남미학과",
    "스포츠과학과" : "스포츠과학과",
    "식품공학과" : "식품공학과",
    "식품영양" : "식품영양학과",
    "신문방송" : "미디어커뮤니케이션학과",
    "미디어커뮤니케이션학" : "미디어커뮤니케이션학과",
    "신소재(금속시스템)" : "신소재공학부(금속시스템공학)",
    "신소재(전자재료)" : "신소재공학부(전자재료공학)",
    "신소재(정보소재)" : "신소재공학부(정보소재공학)",
    "심리" : "심리학과",
    "아동" : "아동학과",
    "약학과" : "약학과",
    "양자시스템공학과" : "양자시스템공학과",
    "역사교육과" : "역사교육과",
    "영어교육" : "영어교육과",
    "영어영문" : "영어영문학과",
    "원예학과" : "원예학과",
    "유기소재섬유공학과" : "유기소재섬유공학과",
    "윤리교육" : "윤리교육과",
    "융합기술공학과" : "융합기술공학부(IT융합기전공학)",
    "융합기술공학부(IT융합자동차공학)" : "융합기술공학부(IT융합기전공학)",
    "융합기술공학부(IT융합기전차공학)" : "융합기술공학부(IT융합기전공학)",
    "융합기술공학부(IT융합기전공학)(재)" : "융합기술공학부(IT융합기전공학)",
    "융합기술공학부(IT융합자동차공학)(재)" : "융합기술공학부(IT융합기전공학)",
    "융합기술공학부(IT응용시스템공학)" : "융합기술공학부(IT응용시스템공학)",
    "융합학부(IAB융합)" : "융합학부",
    "융합학부(반도체 소재/부품/장비)" : "융합학부",
    "융합학부(반도체)" : "융합학부",
    "융합학부(반도체융합)" : "융합학부",
    "융합학부(에너지신산업)" : "융합학부",
    "융합학부(예술융합창작)" : "융합학부",
    "융합학부(메타버스/엔터테인먼트)" : "융합학부",
    "융합학부(방위산업)" : "융합학부",
    "융합학부(지식재산)" : "융합학부",
    "음악과" : "음악과",
    "의" : "의학과",
    "의류" : "의류학과",
    "일반사회교육과" : "일반사회교육과",
    "일본학" : "일본학과",
    "작물생명과학과" : "작물생명과학과",
    "전기공" : "전기공학과",
    "전자공학부" : "전자공학부",
    "정치외교" : "정치외교학과",
    "조경" : "조경학과",
    "주거환경" : "주거환경학과",
    "중어중문" : "중어중문학과",
    "지구환경과학" : "지구환경과학과",
    "지리교육" : "지리교육과",
    "지역건설공학과" : "지역건설공학과",
    "철학과" : "철학과",
    "체육교육" : "체육교육과",
    "치의" : "치의학과",
    "컴퓨터공학부" : "컴퓨터인공지능학부",
    "IT정보공학과" : "컴퓨터인공지능학부",
    "IT지능정보공학과" : "컴퓨터인공지능학부",
    "컴퓨터인공지능학부" : "컴퓨터인공지능학부",
    "컴퓨터공학부(컴퓨터)" : "컴퓨터인공지능학부",
    "토목/환경/자원.에너지공학부(자원.에너지공학)" : "토목환경자원에너지공학부(자원에너지공학)",
    "토목/환경/자원.에너지공학부(토목공학)" : "토목환경자원에너지공학부(토목공학)",
    "토목/환경/자원.에너지공학부(환경공학)" : "토목환경자원에너지공학부(환경공학)",
    "통계학과" : "통계학과",
    "프랑스·아프리카학과" : "프랑스.아프리카학과",
    "한국음악" : "한국음악학과",
    "한약자원학과" : "한약자원학과",
    "한옥학과" : "한옥학과",
    "항공우주공" : "항공우주공학과",
    "행정" : "행정학과",
    "화" : "화학과",
    "화학공학부" : "화학공학부",
    "화공부(나노)" : "화학공학부",
    "화공부(생명)" : "화학공학부",
    "화공부(에너지)" : "화학공학부",
    "회계" : "회계학과",
    "공대" : ["건축공학과", "고분자나노공학과", "유기소재섬유공학과", "기계공학과", "기계설계공학부", "기계시스템공학부", "도시공학과", "바이오메디컬공학부", "산업정보시스템공학과", "소프트웨어공학과", "신소재공학부(금속시스템공학)", "신소재공학부(전자재료공학)", "신소재공학부(정보소재공학)",  "양자시스템공학과", "융합기술공학부(IT융합기전공학)", "융합기술공학부(IT응용시스템공학)", "전기공학과", "전자공학부", "컴퓨터인공지능학부", "토목환경자원에너지공학부(토목공학)", "토목환경자원에너지공학부(환경공학)", "토목환경자원에너지공학부(자원에너지공학)", "항공우주공학과", "화학공학부"],
    "농대" : ["농생물학과", "농경제유통학부", "동물생명공학과", "동물자원과학과", "목재응용과학과", "산림환경과학과", "생명자원융합학과", "생물산업기계공학과", "생물환경화학과", "식품공학과", "원예학과", "작물생명과학과", "조경학과", "지역건설공학과", "농축산식품융합학과"]
}

# 전체 데이터를 이중 리스트로 받음
rows = []


# 5번 column에 해당하는 "학과 학년" 데이터를 학과와 학년으로 분리하여 최종 행 데이터를 rows에 append하는 함수
def department_parse(row):
    department_grade = row[4].rstrip().lstrip()
    grade = department_grade[-1:]
    department = department_grade[:-1].rstrip()
    if grade == "년":
        grade = department_grade[-3]
        department = department_grade[:-3].rstrip()
    
    if grade.isdigit():
        grade = int(grade)
        
        try:
            if isinstance(department_dic[department], list):
                for elements in department_dic[department]:
                    new_row = copy.deepcopy(row)
                    new_row[4] = elements
                    new_row.append(grade)
                    rows.append(new_row)
            else:
                new_row = copy.deepcopy(row)
                new_row[4] = department_dic[department]
                new_row.append(grade)
                rows.append(new_row)
        except:
            print("엑셀에 적힌 학과 이름 : <"+department+">")
        
        
    
# row_value는 한 행의 전체 데이터를 가지는 list
# row_value를 rows에 append 하여 rows는 전체 테이블 데이터를 2중 리스트로 가짐
for i in range(2, ws.max_row+1, 1):
    row_value = []
    
    for j in range(1, 102, 1):
        row_value.append(ws.cell(row=i, column=j).value)
    
    departments = row_value[4].split(',') # 한 강의를 듣는 학과가 여러개인 경우 각각 다른 행으로 처리
    
    if len(departments) > 1:
        for k in departments:
            row_value[4] = k
            department_parse(row_value)
    else :
        department_parse(row_value)
        


# set로 바꿔서 중복 데이터 제거
rows_set = set(tuple(inner_list) for inner_list in rows)


# rows_set 데이터를 mySQL에 insert
for row in rows_set:
    select_query = "select id from jbnu_department where name = %s"
    select_data = (row[4],)
    
    cursor.execute(select_query, select_data)
    department_row = cursor.fetchall()
    try:
        department_id = department_row[0][0]
        print(row)
        long_text_1 = 'YUMMI'.join(row[69:85])
        long_text_2 = 'YUMMI'.join(row[85:101])
        insert_query = "INSERT INTO jbnu_class (offered_year, semester, name, code, student_grade, schedule, section, credits, is_mandatory_for_major, professor, delivery_mode, location, prerequisite, language, description, curriculum, content, department_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        insert_data = (row[0], row[1][0], row[2], row[3], row[101], row[5], row[6], row[7], row[8], row[9], row[10], row[12], row[25], row[32], row[13], long_text_1, long_text_2, department_id)
        # 학년은 제일 마지막에 들어감 조심해야함
        cursor.execute(insert_query, insert_data)
        connection.commit()
    except Exception as e:
        print(e)
        print("실제 학과 이름 : <"+row[4]+">")
        
    

connection.close()
